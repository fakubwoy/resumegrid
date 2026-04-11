from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
import os
import json
import re
import subprocess
import tempfile
import time
import logging
import base64
import pdfplumber
import pypdf
import docx2txt
import httpx
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# OCR imports — optional, gracefully degraded if unavailable
try:
    from pdf2image import convert_from_path as pdf_to_images
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Logging setup ──────────────────────────────────────────────────────────────

LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("resumegrid")

logging.getLogger("pdfplumber").setLevel(logging.WARNING)
logging.getLogger("pdfminer").setLevel(logging.WARNING)
logging.getLogger("pypdf").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)

# ── Provider configuration — Gemini only ──────────────────────────────────────

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL   = os.environ.get("GEMINI_MODEL", "gemini-2.0-flash")

PRIMARY_PROVIDER       = "gemini"
MAX_CONCURRENT_AI      = int(os.environ.get("MAX_CONCURRENT_AI", "20"))
MAX_CONCURRENT_EXTRACT = int(os.environ.get("MAX_CONCURRENT_EXTRACT", "8"))

logger.info("ResumeGrid starting up (log level: %s)", LOG_LEVEL)
logger.info("AI provider: gemini (%s) | concurrent calls: %d", GEMINI_MODEL, MAX_CONCURRENT_AI)
if OCR_AVAILABLE:
    logger.info("OCR support enabled (pdf2image + pytesseract)")
else:
    logger.warning("OCR support NOT available — install pdf2image + pytesseract to enable.")

if not GEMINI_API_KEY:
    logger.warning("GEMINI_API_KEY not set — all AI calls will fail.")


# ── AI semaphore ──────────────────────────────────────────────────────────────

AI_SEMAPHORE = threading.Semaphore(MAX_CONCURRENT_AI)


# ── Gemini / call_ai ─────────────────────────────────────────────────────────

def call_ai(text, is_scoring=False):
    """
    Call Gemini. Retries once on transient errors, raises on 429.
    Returns (raw_response_str, "gemini").
    """
    system = (
        "You are an expert resume parser. "
        "Return ONLY valid JSON. No markdown, no backticks, no explanation."
    ) if not is_scoring else (
        "You are a technical recruiter. Return ONLY valid JSON. No markdown, no explanation."
    )

    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": text}]}],
        "generationConfig": {
            "temperature": 0.0,
            "maxOutputTokens": 2048,
            "responseMimeType": "application/json",
        },
        "system_instruction": {"parts": [{"text": system}]},
    }

    last_err = None
    for attempt in range(2):
        try:
            with AI_SEMAPHORE:
                resp = httpx.post(url, json=payload, timeout=45.0)

            if resp.status_code == 429:
                raise RuntimeError("Gemini 429 — quota exceeded. Try again in a moment.")

            resp.raise_for_status()
            return resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip(), "gemini"

        except Exception as e:
            last_err = e
            if attempt == 0:
                time.sleep(1)

    raise RuntimeError(f"Gemini failed: {last_err}")

# ──────────────────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder='static')
CORS(app)

UPLOAD_FOLDER = tempfile.mkdtemp()

CANONICAL_FIELDS = {
    "full_name": ["name", "full name", "candidate name", "applicant name", "your name"],
    "email": ["email", "email address", "e-mail", "mail", "contact email"],
    "phone": ["phone", "phone number", "mobile", "cell", "telephone", "contact number", "tel"],
    "location": ["location", "address", "city", "city/state", "current location", "residence", "based in"],
    "linkedin": ["linkedin", "linkedin url", "linkedin profile", "linkedin.com"],
    "github": ["github", "github url", "github profile", "github.com"],
    "portfolio": ["portfolio", "website", "personal website", "portfolio url"],
    "current_title": ["current title", "job title", "position", "current position", "role", "current role"],
    "years_of_experience": ["years of experience", "experience", "total experience", "work experience years"],
    "summary": ["summary", "objective", "profile", "about", "professional summary", "career objective", "about me"],
    "skills": ["skills", "technical skills", "core skills", "key skills", "competencies", "technologies", "tech stack", "tools"],
    "programming_languages": ["programming languages", "languages", "coding languages", "development languages"],
    "frameworks": ["frameworks", "libraries", "frameworks & libraries", "tools & frameworks"],
    "education_degree": ["degree", "education", "highest degree", "qualification", "academic background"],
    "education_institution": ["university", "college", "institution", "school", "alma mater"],
    "education_year": ["graduation year", "year of graduation", "year of passing", "completion year"],
    "education_gpa": ["gpa", "cgpa", "grade", "percentage", "marks"],
    "companies_worked": ["companies", "employers", "work history", "previous companies", "organisations"],
    "most_recent_company": ["most recent company", "current company", "last employer", "current employer"],
    "most_recent_role": ["most recent role", "current role", "latest position", "last position"],
    "most_recent_duration": ["most recent duration", "current duration", "last job duration"],
    "certifications": ["certifications", "certificates", "professional certifications", "credentials", "licenses"],
    "languages_spoken": ["languages spoken", "languages known", "spoken languages", "language proficiency"],
    "projects": ["projects", "key projects", "notable projects", "personal projects", "academic projects"],
    "achievements": ["achievements", "accomplishments", "awards", "honors", "recognition"],
    "total_companies": ["total companies", "number of companies", "companies count"],
    "last_ctc": ["last ctc", "current ctc", "last salary", "current salary", "compensation", "ctc"],
    "current_status": ["current status", "employment status", "work status", "availability"]
}

COLUMN_ORDER = [
    "match_score", "match_reason", "skill_depth",
    "full_name", "email", "phone", "location", "linkedin", "github", "portfolio",
    "current_title", "years_of_experience", "summary",
    "skills", "programming_languages", "frameworks",
    "education_degree", "education_institution", "education_year", "education_gpa",
    "companies_worked", "most_recent_company", "most_recent_role", "most_recent_duration", "total_companies",
    "certifications", "languages_spoken", "projects", "achievements",
    "last_ctc", "current_status", "duplicate_of"
]

COLUMN_HEADERS = {
    "full_name": "Full Name", "email": "Email", "phone": "Phone",
    "location": "Location", "linkedin": "LinkedIn", "github": "GitHub",
    "portfolio": "Portfolio/Website", "current_title": "Current Title",
    "years_of_experience": "Years of Experience", "summary": "Summary/Objective",
    "skills": "Skills", "programming_languages": "Programming Languages",
    "frameworks": "Frameworks/Libraries", "education_degree": "Degree",
    "education_institution": "Institution", "education_year": "Graduation Year",
    "education_gpa": "GPA/Grade", "companies_worked": "Companies Worked At",
    "most_recent_company": "Most Recent Company", "most_recent_role": "Most Recent Role",
    "most_recent_duration": "Most Recent Duration", "total_companies": "Total Companies",
    "certifications": "Certifications", "languages_spoken": "Languages Spoken",
    "projects": "Key Projects", "achievements": "Achievements/Awards",
    "source_file": "Source File", "last_ctc": "Last CTC", "current_status": "Current Status",
    "match_score": "Match Score (/100)", "match_reason": "Match Reason", "skill_depth": "Skill Depth (per JD)",
    "duplicate_of": "Duplicate Of"
}


# ── Text extraction helpers ───────────────────────────────────────────────────

def extract_hyperlinks_from_pdf(tmp_path):
    urls = []
    try:
        reader = pypdf.PdfReader(tmp_path)
        for page in reader.pages:
            annots = page.get("/Annots")
            if not annots:
                continue
            for annot in annots:
                obj = annot.get_object()
                a = obj.get("/A")
                if not a:
                    continue
                a_obj = a.get_object() if hasattr(a, 'get_object') else a
                uri = a_obj.get("/URI")
                if uri:
                    url = str(uri)
                    if url.startswith("http") and url not in urls:
                        urls.append(url)
    except Exception:
        pass
    return urls


def is_image_based_pdf(text, page_count):
    stripped = text.strip() if text else ""
    if len(stripped) < page_count * 300:
        return True
    import re as _re
    cid_hits = len(_re.findall(r'\(cid:\d+\)', stripped))
    word_count = len(stripped.split())
    if word_count > 0 and cid_hits / word_count > 0.10:
        return True
    return False


def ocr_pdf(tmp_path, filename):
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "OCR is not available. Install pdf2image and pytesseract to process image-based PDFs."
        )
    logger.info("'%s' — running OCR (image-based PDF detected)", filename)
    t0 = time.monotonic()
    poppler_path = os.environ.get("POPPLER_PATH", "/usr/bin")
    pages = pdf_to_images(tmp_path, dpi=200, poppler_path=poppler_path)
    parts = []
    for i, page_img in enumerate(pages):
        page_text = pytesseract.image_to_string(page_img, lang="eng")
        if page_text.strip():
            parts.append(page_text)
    text = "\n".join(parts)
    elapsed = time.monotonic() - t0
    logger.info("'%s' — OCR complete in %.2fs: %d pages → %d chars", filename, elapsed, len(pages), len(text))
    return text


def extract_text_via_gemini_vision(tmp_path, filename):
    """
    Use Gemini Vision to extract text from image-based PDFs.

    Strategy (in order of preference):
    1. Send the PDF directly as inline_data with mime_type application/pdf —
       Gemini 1.5+ natively reads PDFs without poppler or any conversion.
    2. If pdf_to_images (poppler) IS available, fall back to per-page JPEG method.

    This means poppler is completely optional — we only use it if Gemini PDF
    inlining somehow fails.
    """
    import base64 as _b64
    import io as _io

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    extraction_prompt = (
        "This is a resume/CV document. Extract ALL text content exactly as it appears, "
        "preserving the reading order and structure. Include every piece of information: "
        "name, contact details, work experience, education, skills, projects, certifications, "
        "achievements, etc. Output only the extracted text, nothing else."
    )

    # ── Strategy 1: Send PDF directly to Gemini (no poppler needed) ──────────
    try:
        with open(tmp_path, "rb") as f:
            pdf_bytes = f.read()
        pdf_b64 = _b64.b64encode(pdf_bytes).decode("utf-8")

        payload = {
            "contents": [{
                "role": "user",
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": "application/pdf",
                            "data": pdf_b64
                        }
                    },
                    {"text": extraction_prompt}
                ]
            }],
            "generationConfig": {"temperature": 0.0, "maxOutputTokens": 8192}
        }

        resp = httpx.post(url, json=payload, timeout=90.0)
        if resp.status_code == 429:
            raise RateLimitError("Gemini Vision 429")
        resp.raise_for_status()
        result = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        if result and len(result) > 100:
            logger.info("'%s' — Gemini PDF-inline extracted %d chars", filename, len(result))
            return result
        logger.warning("'%s' — Gemini PDF-inline returned short text (%d chars), trying page-image fallback",
                       filename, len(result))
    except RateLimitError:
        raise
    except Exception as e:
        logger.warning("'%s' — Gemini PDF-inline failed (%s), trying page-image fallback", filename, e)

    # ── Strategy 2: Per-page JPEG (requires pdf2image + poppler) ─────────────
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "Gemini PDF-inline failed and pdf2image/poppler is not installed. "
            "Cannot process image-based PDF without one of these."
        )

    poppler_path = os.environ.get("POPPLER_PATH", "/usr/bin")
    pages = pdf_to_images(tmp_path, dpi=200, poppler_path=poppler_path)
    all_text_parts = []

    for i, page_img in enumerate(pages):
        buf = _io.BytesIO()
        page_img.save(buf, format="JPEG", quality=90)
        img_b64 = _b64.b64encode(buf.getvalue()).decode("utf-8")

        payload = {
            "contents": [{
                "role": "user",
                "parts": [
                    {"inline_data": {"mime_type": "image/jpeg", "data": img_b64}},
                    {"text": (
                        "This is page of a resume. Extract ALL text exactly as it appears, "
                        "preserving structure. Output only the extracted text, nothing else."
                    )}
                ]
            }],
            "generationConfig": {"temperature": 0.0, "maxOutputTokens": 4096}
        }

        resp = httpx.post(url, json=payload, timeout=60.0)
        if resp.status_code == 429:
            raise RateLimitError("Gemini Vision 429")
        resp.raise_for_status()
        page_text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
        if page_text:
            all_text_parts.append(page_text)

    result = "\n\n".join(all_text_parts)
    logger.info("'%s' — Gemini page-image extracted %d chars from %d page(s)", filename, len(result), len(pages))
    return result


def classify_urls(urls):
    result = {}
    for url in urls:
        url_lower = url.lower()
        if "linkedin.com" in url_lower and "linkedin" not in result:
            result["linkedin"] = url
        elif "github.com" in url_lower and "github" not in result:
            result["github"] = url
        elif "mailto:" in url_lower and "email" not in result:
            result["email"] = url.replace("mailto:", "").strip()
        elif "portfolio" not in result and "linkedin.com" not in url_lower and "github.com" not in url_lower:
            result["portfolio"] = url
    return result


def extract_text_from_pdf_smart(tmp_path, filename):
    """
    Multi-strategy PDF text extraction:
    1. pdfplumber with layout=True (preserves column order better than default)
    2. Falls back to pdfplumber default if layout mode gives less text
    Returns (text, num_pages, is_image_based)
    """
    best_text = ""
    num_pages = 1
    try:
        with pdfplumber.open(tmp_path) as pdf:
            num_pages = len(pdf.pages)

            # Strategy A: layout=True — better for multi-column, preserves spatial reading order
            parts_layout = []
            for page in pdf.pages:
                try:
                    t = page.extract_text(layout=True)
                    if t:
                        parts_layout.append(t)
                except Exception:
                    pass
            text_layout = "\n".join(parts_layout)

            # Strategy B: default extraction — sometimes better for simple single-column
            parts_default = []
            for page in pdf.pages:
                try:
                    t = page.extract_text()
                    if t:
                        parts_default.append(t)
                except Exception:
                    pass
            text_default = "\n".join(parts_default)

            # Pick whichever got more real content (fewer cid: artifacts)
            def quality_score(t):
                import re as _re
                if not t:
                    return 0
                cid_hits = len(_re.findall(r'\(cid:\d+\)', t))
                return len(t.strip()) - cid_hits * 5

            best_text = text_layout if quality_score(text_layout) >= quality_score(text_default) else text_default

    except Exception as e:
        logger.warning("'%s' — pdfplumber failed: %s", filename, e)

    image_based = is_image_based_pdf(best_text, num_pages)
    return best_text, num_pages, image_based


def extract_text_from_file(file_content, filename):
    ext = Path(filename).suffix.lower()
    logger.debug("Extracting text from '%s' (type: %s, size: %d bytes)", filename, ext, len(file_content))

    if ext == ".pdf":
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name
        try:
            text, num_pages, image_based = extract_text_from_pdf_smart(tmp_path, filename)
            urls = extract_hyperlinks_from_pdf(tmp_path)
            url_overrides = classify_urls(urls)

            if image_based:
                # Try Gemini Vision first (best accuracy for complex layouts/handwriting)
                # Falls back to Tesseract OCR if Gemini not available
                logger.info("'%s' — image-based PDF detected, trying vision extraction", filename)
                vision_text = None

                if GEMINI_API_KEY and GEMINI_API_KEY.strip():
                    try:
                        vision_text = extract_text_via_gemini_vision(tmp_path, filename)
                    except Exception as ve:
                        logger.warning("'%s' — Gemini Vision failed, falling back to OCR: %s", filename, ve)

                if not vision_text and OCR_AVAILABLE:
                    try:
                        vision_text = ocr_pdf(tmp_path, filename)
                    except Exception as ocr_err:
                        logger.error("'%s' — OCR failed: %s", filename, ocr_err)

                if vision_text and len(vision_text.strip()) > len(text.strip()):
                    text = vision_text

            return text, url_overrides
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    elif ext in [".doc", ".docx"]:
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(file_content)
            tmp_path = tmp.name
        try:
            text = docx2txt.process(tmp_path)
            url_overrides = extract_hyperlinks_from_docx(tmp_path)
            return text, url_overrides
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    logger.warning("'%s' — falling back to plain text decode", filename)
    return file_content.decode("utf-8", errors="ignore"), {}


def extract_hyperlinks_from_docx(tmp_path):
    urls = []
    try:
        from docx import Document
        doc = Document(tmp_path)
        for rel in doc.part.rels.values():
            if "hyperlink" in rel.reltype.lower():
                url = rel.target_ref
                if url.startswith("http") and url not in urls:
                    urls.append(url)
    except Exception:
        pass
    return classify_urls(urls)


# ── AI prompts ─────────────────────────────────────────────────────────────────

def get_extraction_prompt():
    return """Extract ALL information from this resume. Return a single JSON object.

Fields to extract (use null if genuinely not present):
- full_name: candidate's full name
- email, phone, location
- linkedin: full LinkedIn URL or handle
- github: full GitHub URL or handle
- portfolio: personal website/portfolio URL
- current_title: their current or most recent job title
- years_of_experience: total years (compute from dates if not stated explicitly)
- summary: professional summary or objective paragraph
- skills: all technical and soft skills, comma-separated
- programming_languages: programming languages only, comma-separated
- frameworks: frameworks, libraries, tools, comma-separated
- education_degree: highest degree (e.g. "B.Tech Computer Science", "MBA Marketing")
- education_institution: university/college name
- education_year: graduation year or expected year
- education_gpa: GPA, CGPA, or percentage if mentioned
- companies_worked: ALL companies/organisations ever worked at, comma-separated (include internships)
- most_recent_company: name of their most recent employer
- most_recent_role: title at most recent company
- most_recent_duration: duration at most recent company (e.g. "Jun 2024 - Present", "2 years")
- total_companies: count of distinct companies/organisations
- certifications: all certifications and courses, comma-separated
- languages_spoken: human languages (English, Hindi, etc.), comma-separated
- projects: key project names and one-line descriptions, semicolon-separated
- achievements: awards, honours, notable accomplishments
- last_ctc: stated salary/CTC/compensation — null if not mentioned
- current_status: MUST be exactly one of "Employed", "Not Employed", or "Fresher"

Rules for current_status:
- "Employed": has a current role, or dates show "Present" / ongoing
- "Not Employed": most recent role ended and no current role shown
- "Fresher": no professional work experience (internships alone = Fresher)

Rules for experience fields:
- Look for EXPERIENCE, WORK HISTORY, EMPLOYMENT sections carefully
- Internships count as experience for companies_worked and total_companies
- If dates are given (e.g. "Jun 2023 - Jul 2023"), calculate duration yourself
- years_of_experience: sum all work durations; if only internships and no full-time, state "< 1 year"

CRITICAL: Return ONLY a valid JSON object. No markdown, no backticks, no explanation."""


def get_scoring_prompt(job_description, applied_role=None):
    role_instruction = ""
    if applied_role:
        role_instruction = f"""
CANDIDATE'S APPLIED ROLE: {applied_role}

CRITICAL: If the candidate applied for a role that is clearly different from the role described in the JD above (e.g., they applied for "Full Stack Developer" but this JD is for "Entrepreneur in Residence"), cap the score at 20 and note the mismatch in your reason. Only score 21-100 if the applied role is reasonably aligned with this JD.
"""
    return f"""You are a senior technical recruiter evaluating a candidate against a job description.

JOB DESCRIPTION:
{job_description[:4000]}
{role_instruction}

SCORING INSTRUCTIONS:

Step 1 - Score the candidate 0-100:
- Role alignment: is the candidate's applied role relevant to this JD? (if not, cap at 20)
- Skills & technology match (35 points of 80 remaining)
- Years of experience relevance (20 points)
- Role/title alignment (15 points)
- Education & certifications (10 points)

Step 2 - For each skill/technology explicitly required or strongly implied by the JD, infer the
candidate's depth from HOW they describe using it in the resume, not just whether the word appears.

Depth levels:
- "expert"       : led projects, architected systems, mentored others, 3+ years direct use, or measurable impact described
- "intermediate" : used independently on real projects, 1-3 years, clear hands-on usage described
- "beginner"     : mentioned briefly, listed without context, coursework/academic only, or < 1 year
- "not found"    : not mentioned or no evidence at all

Be strict. "Familiar with X" or listing X in a skills section with no project evidence = beginner at best.
If a skill is not mentioned at all, mark it "not found" - do not assume.

Return ONLY a JSON object with exactly these three fields:
{{
  "match_score": <integer 0-100>,
  "match_reason": "<2-3 sentence summary: overall fit, strongest signals, and any key gaps>",
  "skill_depth": {{
    "<skill_name>": "<expert|intermediate|beginner|not found>",
    "...more skills...": "..."
  }}
}}

Include only the top 8-12 most important skills from the JD in skill_depth.
No markdown, no explanation, just the JSON."""


# ── Core extraction ───────────────────────────────────────────────────────────

def normalize_fields(data):
    normalized = {}
    for key, value in data.items():
        if value is None or value == "" or value == "null":
            continue
        key_lower = key.lower().strip()
        if key_lower in CANONICAL_FIELDS:
            normalized[key_lower] = str(value)
            continue
        matched = False
        for canonical, aliases in CANONICAL_FIELDS.items():
            if key_lower in aliases or any(alias in key_lower for alias in aliases):
                normalized[canonical] = str(value)
                matched = True
                break
        if not matched:
            normalized[key_lower] = str(value)
    return normalized


def extract_single_resume(filename, text, url_overrides, job_description=None):
    """
    Full AI extraction + optional JD scoring for one resume.
    This entire function runs in a thread — fully parallel across resumes.
    """
    t0 = time.monotonic()
    logger.info("AI extracting '%s'", filename)

    prompt = f"Resume text:\n\n{text[:20000]}\n\n{get_extraction_prompt()}"
    raw, provider_used = call_ai(prompt)

    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)

    match_obj = re.search(r'\{.*\}', raw, re.DOTALL)
    data = {}
    if match_obj:
        try:
            data = normalize_fields(json.loads(match_obj.group()))
        except json.JSONDecodeError as e:
            logger.warning("'%s' — JSON parse error: %s", filename, e)
    else:
        logger.warning("'%s' — no JSON in AI response", filename)

    for field, url in (url_overrides or {}).items():
        existing = data.get(field, "")
        if not existing or not existing.startswith("http"):
            data[field] = url

    data["filename"] = filename

    # JD scoring (also via AI, fully parallel)
    if job_description:
        try:
            # Build rich context: extracted fields + full raw resume text.
            # This ensures the scorer sees everything — project descriptions, impact
            # statements, work history details — not just a 10-field skeleton.
            extracted_fields = []
            for field in ["full_name", "current_title", "years_of_experience", "skills",
                          "programming_languages", "frameworks", "education_degree",
                          "certifications", "summary", "companies_worked",
                          "most_recent_role", "most_recent_company", "most_recent_duration",
                          "projects", "achievements", "languages_spoken"]:
                val = data.get(field)
                if val:
                    extracted_fields.append(f"{field}: {val}")
            extracted_summary = "\n".join(extracted_fields)
            # Full raw resume text (capped at 15k) so no detail is lost during scoring
            full_resume_context = (
                f"--- EXTRACTED FIELDS ---\n{extracted_summary}\n\n"
                f"--- FULL RESUME TEXT ---\n{text[:15000]}"
            )
            # Pass the candidate's current_title as the "applied role" so the model
            # can detect role mismatches (e.g. a fullstack dev ranked against an EIR JD)
            applied_role = data.get("current_title") or data.get("most_recent_role")
            score_prompt = f"RESUME DATA:\n{full_resume_context}\n\n{get_scoring_prompt(job_description, applied_role=applied_role)}"
            score_raw, _ = call_ai(score_prompt, is_scoring=True)
            score_raw = re.sub(r"^```[a-z]*\n?", "", score_raw)
            score_raw = re.sub(r"\n?```$", "", score_raw)
            score_match = re.search(r'\{.*\}', score_raw, re.DOTALL)
            if score_match:
                parsed = json.loads(score_match.group())
                score = max(0, min(100, int(parsed.get("match_score", 0))))
                data["match_score"] = str(score)
                data["match_reason"] = str(parsed.get("match_reason", ""))
                # Serialize skill_depth dict into a readable string for Excel/grid
                skill_depth = parsed.get("skill_depth", {})
                if isinstance(skill_depth, dict) and skill_depth:
                    depth_lines = []
                    order = ["expert", "intermediate", "beginner", "not found"]
                    for level in order:
                        skills_at_level = [k for k, v in skill_depth.items() if str(v).lower() == level]
                        if skills_at_level:
                            emoji = {"expert": "\u2605\u2605\u2605", "intermediate": "\u2605\u2605\u2606", "beginner": "\u2605\u2606\u2606", "not found": "\u2717"}[level]
                            depth_lines.append(f"{emoji} {level.title()}: {', '.join(skills_at_level)}")
                    data["skill_depth"] = "\n".join(depth_lines)
                logger.info("'%s' \u2014 match score: %s/100 | skill_depth: %d skills rated", filename, score, len(skill_depth))
        except Exception as score_err:
            logger.warning("Scoring skipped for '%s': %s", filename, score_err)

    elapsed = time.monotonic() - t0
    fields_found = [k for k in data if data[k] and k != "filename"]
    logger.info("'%s' — OK in %.2fs via %s (%d fields)", filename, elapsed, provider_used, len(fields_found))

    return data


# ── Excel generation ──────────────────────────────────────────────────────────

def create_excel(all_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    all_keys = set()
    for d in all_data:
        all_keys.update(d.keys())

    final_columns = [col for col in COLUMN_ORDER if col in all_keys or col in COLUMN_HEADERS]
    for key in sorted(all_keys):
        if key not in final_columns and key != "filename":
            final_columns.append(key)
    final_columns.append("source_file")

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1A3A5C")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font    = Font(name="Arial", size=9)
    cell_align   = Alignment(vertical="top", wrap_text=True)
    alt_fill     = PatternFill("solid", start_color="EBF2FA")
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx, col_key in enumerate(final_columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = COLUMN_HEADERS.get(col_key, col_key.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    for row_idx, candidate in enumerate(all_data, 2):
        for col_idx, col_key in enumerate(final_columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = candidate.get("filename", "") if col_key == "source_file" else candidate.get(col_key, "")

            if value and str(value).startswith("http") and col_key in ("linkedin", "github", "portfolio"):
                cell.value = value
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            elif col_key == "match_score" and value:
                try:
                    score_val = int(value)
                    cell.value = score_val
                    if score_val >= 75:
                        cell.fill = PatternFill("solid", start_color="C6EFCE")
                        cell.font = Font(name="Arial", size=9, bold=True, color="276221")
                    elif score_val >= 50:
                        cell.fill = PatternFill("solid", start_color="FFEB9C")
                        cell.font = Font(name="Arial", size=9, bold=True, color="9C6500")
                    else:
                        cell.fill = PatternFill("solid", start_color="FFC7CE")
                        cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")
                except (ValueError, TypeError):
                    cell.value = value
                    cell.font = cell_font
            elif col_key == "duplicate_of" and value:
                cell.value = value
                cell.fill = PatternFill("solid", start_color="FFF2CC")
                cell.font = Font(name="Arial", size=9, italic=True, color="7F6000")
            else:
                cell.value = value
                cell.font = cell_font

            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0 and not (value and str(value).startswith("http") and col_key in ("linkedin", "github", "portfolio")):
                cell.fill = alt_fill

    width_map = {
        "full_name": 22, "email": 28, "phone": 16, "location": 18,
        "linkedin": 35, "github": 35, "portfolio": 35, "current_title": 22,
        "years_of_experience": 12, "summary": 45, "skills": 40,
        "programming_languages": 30, "frameworks": 30, "education_degree": 20,
        "education_institution": 25, "education_year": 12, "education_gpa": 10,
        "companies_worked": 35, "most_recent_company": 25, "most_recent_role": 22,
        "most_recent_duration": 15, "total_companies": 12, "certifications": 35,
        "languages_spoken": 20, "projects": 45, "achievements": 35, "source_file": 25,
        "last_ctc": 18, "current_status": 18, "match_score": 14,
        "match_reason": 45, "duplicate_of": 25
    }
    for col_idx, col_key in enumerate(final_columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(col_key, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(final_columns))}1"

    output_path = os.path.join(UPLOAD_FOLDER, "extracted_resumes.xlsx")
    wb.save(output_path)
    with open(output_path, "rb") as f:
        excel_bytes = f.read()
    return output_path, excel_bytes


# ── Flask routes ──────────────────────────────────────────────────────────────

def sse_event(data):
    return f"data: {json.dumps(data)}\n\n"


@app.route("/")
def index():
    return app.send_static_file('index.html')


@app.route("/health")
def health():
    """Health check."""
    return jsonify({
        "status": "ok",
        "provider": "gemini",
        "model": GEMINI_MODEL,
        "gemini_configured": bool(GEMINI_API_KEY),
        "max_concurrent_ai": MAX_CONCURRENT_AI,
    })


@app.route("/extract", methods=["POST"])
def extract():
    """
    Streaming SSE endpoint — fully parallel AI extraction.

    Architecture:
    - Phase 1: Text extraction from all files in parallel (ThreadPoolExecutor)
    - Phase 2: AI extraction for all files in parallel (ThreadPoolExecutor)
      - With Gemini/Groq: 20+ concurrent calls → massive speedup
      - With Ollama: serialized via semaphore (GPU constraint)
    - Results streamed back as SSE as each resume completes
    """
    if "files" not in request.files:
        return jsonify({"error": "No files uploaded"}), 400

    files = request.files.getlist("files")
    job_description = request.form.get("job_description", "").strip()
    if job_description:
        logger.info("Job description provided (%d chars) — scoring enabled", len(job_description))

    file_payloads = []
    for file in files:
        if not file.filename:
            continue
        ext = Path(file.filename).suffix.lower()
        if ext not in [".pdf", ".doc", ".docx"]:
            file_payloads.append({"filename": file.filename, "content": None,
                                   "error": "Unsupported file type"})
        else:
            content = file.read()
            file_payloads.append({"filename": file.filename, "content": content, "error": None})

    if not file_payloads:
        return jsonify({"error": "No valid files found"}), 400

    total = len(file_payloads)
    logger.info("POST /extract — starting batch of %d file(s) | provider: %s | concurrency: %d",
                total, PRIMARY_PROVIDER, MAX_CONCURRENT_AI)

    def generate():
        results = []
        errors  = []
        batch_start = time.monotonic()
        seen_emails = {}
        seen_phones = {}
        state_lock = threading.Lock()
        completed_count = [0]  # mutable for closure

        yield sse_event({"type": "start", "total": total})

        # ── Phase 1: extract text from all files in parallel ──────────────────
        text_results = {}

        def extract_text_task(payload):
            fn = payload["filename"]
            if payload["error"]:
                return fn, None, None, payload["error"]
            try:
                text, url_overrides = extract_text_from_file(payload["content"], fn)
                if not text or len(text.strip()) < 50:
                    return fn, None, None, "Could not extract readable text from file"
                return fn, text, url_overrides, None
            except Exception as e:
                return fn, None, None, str(e)

        logger.info("Phase 1: extracting text from %d files (up to %d in parallel)", total, MAX_CONCURRENT_EXTRACT)
        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_EXTRACT) as executor:
            futures = {executor.submit(extract_text_task, p): p for p in file_payloads}
            for future in as_completed(futures):
                fn, text, url_overrides, err = future.result()
                text_results[fn] = (text, url_overrides, err)

        # ── Phase 2: AI extraction — ALL files in parallel ────────────────────
        logger.info("Phase 2: AI extraction for %d files (%d concurrent, provider: %s)",
                    total, MAX_CONCURRENT_AI, PRIMARY_PROVIDER)

        # Queue of SSE events to stream (thread-safe)
        event_queue = []
        event_lock = threading.Lock()

        def process_one(payload):
            filename = payload["filename"]
            text, url_overrides, text_err = text_results.get(filename, (None, None, "Text extraction missing"))

            with state_lock:
                completed_count[0] += 1
                completed = completed_count[0]

            if text_err:
                logger.warning("[%d/%d] Skipping '%s': %s", completed, total, filename, text_err)
                errors.append({"file": filename, "error": text_err})
                return {"type": "progress", "filename": filename, "ok": False,
                        "error": text_err, "completed": completed, "total": total,
                        "pct": round((completed / total) * 90)}

            try:
                data = extract_single_resume(filename, text, url_overrides, job_description)

                # Duplicate detection (needs lock since parallel)
                dup_of = None
                email_key = (data.get("email") or "").lower().strip()
                phone_key  = re.sub(r'\D', '', data.get("phone") or "")

                with state_lock:
                    if email_key and email_key in seen_emails:
                        dup_of = seen_emails[email_key]
                    elif phone_key and len(phone_key) >= 7 and phone_key in seen_phones:
                        dup_of = seen_phones[phone_key]

                    if dup_of:
                        data["duplicate_of"] = dup_of
                    else:
                        candidate_label = data.get("full_name") or filename
                        if email_key:
                            seen_emails[email_key] = candidate_label
                        if phone_key and len(phone_key) >= 7:
                            seen_phones[phone_key] = candidate_label

                    results.append(data)

                return {"type": "progress", "filename": filename, "ok": True,
                        "error": None, "completed": completed, "total": total,
                        "pct": round((completed / total) * 90), "data": data}

            except Exception as e:
                logger.error("[%d/%d] '%s' — FAILED: %s", completed, total, filename, e, exc_info=True)
                with state_lock:
                    errors.append({"file": filename, "error": str(e)})
                return {"type": "progress", "filename": filename, "ok": False,
                        "error": str(e), "completed": completed, "total": total,
                        "pct": round((completed / total) * 90)}

        # Run all AI extractions in parallel, yield SSE as each completes
        with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_AI) as executor:
            future_to_payload = {executor.submit(process_one, p): p for p in file_payloads}
            for future in as_completed(future_to_payload):
                try:
                    event = future.result()
                    yield sse_event(event)
                except Exception as e:
                    payload = future_to_payload[future]
                    logger.error("Unexpected error for '%s': %s", payload["filename"], e)
                    yield sse_event({"type": "progress", "filename": payload["filename"],
                                     "ok": False, "error": str(e),
                                     "completed": total, "total": total, "pct": 90})

        batch_elapsed = time.monotonic() - batch_start
        logger.info("Batch complete in %.2fs — %d/%d succeeded, %d error(s)",
                    batch_elapsed, len(results), total, len(errors))

        if results:
            try:
                _, excel_bytes = create_excel(results)
                excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")
                yield sse_event({
                    "type": "done", "success": True,
                    "processed": len(results), "errors": errors,
                    "download_url": "/download", "pct": 100,
                    "results": results, "excel_b64": excel_b64
                })
            except Exception as e:
                logger.error("Excel generation failed: %s", e, exc_info=True)
                yield sse_event({
                    "type": "done", "success": False,
                    "error": f"Excel generation failed: {str(e)}",
                    "processed": 0, "errors": errors, "pct": 0
                })
        else:
            yield sse_event({
                "type": "done", "success": False,
                "error": "No resumes could be processed",
                "processed": 0, "errors": errors, "pct": 0
            })

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.route("/download")
def download():
    excel_path = os.path.join(UPLOAD_FOLDER, "extracted_resumes.xlsx")
    if not os.path.exists(excel_path):
        return jsonify({"error": "No file to download"}), 404
    return send_file(
        excel_path, as_attachment=True,
        download_name="extracted_resumes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# CANDIDATE GRID — Routes for Google Form / sheet candidate review
# ═══════════════════════════════════════════════════════════════════════════════

import urllib.parse as _urlparse
import urllib.request as _gdrive_req


@app.route("/candidates")
def candidate_grid():
    """Serve the candidate grid page."""
    return app.send_static_file("candidate_grid.html")


def _gdrive_direct_url(url):
    """
    Convert a Google Drive share link → direct download URL.
    Handles /file/d/<ID>/view and ?id=<ID> patterns.
    """
    m = re.search(r"/d/([a-zA-Z0-9_-]{20,})", url)
    if not m:
        m = re.search(r"[?&]id=([a-zA-Z0-9_-]{20,})", url)
    if m:
        file_id = m.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    return url


def _is_gdrive_folder(url):
    """Return True if the URL points to a Google Drive folder."""
    return bool(re.search(r"drive\.google\.com/drive/folders/", url))


def _gdrive_folder_id(url):
    """Extract the folder ID from a Google Drive folder URL."""
    m = re.search(r"/folders/([a-zA-Z0-9_-]{20,})", url)
    return m.group(1) if m else None


def _find_resume_in_gdrive_folder(folder_id, timeout=12):
    """
    Optimized: parallel probing of candidate file IDs to speed up folder crawling.
    """
    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }

    try:
        req = _gdrive_req.Request(folder_url, headers=headers)
        with _gdrive_req.urlopen(req, timeout=timeout) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as exc:
        logger.warning("Could not fetch folder page %s: %s", folder_id, exc)
        return None

    # Extract candidate file IDs
    candidate_ids = []
    for m in re.finditer(r'["\\/]d["\\/]([a-zA-Z0-9_-]{25,})["\\/]', html):
        fid = m.group(1)
        if fid not in candidate_ids and fid != folder_id:
            candidate_ids.append(fid)
    for m in re.finditer(r'\b([a-zA-Z0-9_-]{33})\b', html):
        fid = m.group(1)
        if fid not in candidate_ids and fid != folder_id:
            candidate_ids.append(fid)

    if not candidate_ids:
        logger.info("No file IDs found in folder %s HTML", folder_id)
        return None

    logger.info("Found %d candidate file IDs in folder %s — parallel probing for PDF/DOCX",
                len(candidate_ids), folder_id)

    # Limit to first 15 IDs to keep reasonable
    ids_to_probe = candidate_ids[:15]

    def probe_file_id(fid):
        """Check if a file ID is a PDF or DOCX resume."""
        probe_url = f"https://drive.google.com/uc?export=download&id={fid}"
        try:
            # Try HEAD first (lightweight)
            head_req = _gdrive_req.Request(probe_url, headers=headers, method="HEAD")
            with _gdrive_req.urlopen(head_req, timeout=5) as resp:
                ct = resp.headers.get("Content-Type", "").lower()
                cl = int(resp.headers.get("Content-Length", "0") or "0")
            if cl > 5000 and ("pdf" in ct or "word" in ct or "docx" in ct or "zip" in ct):
                return fid, probe_url
        except Exception:
            pass

        # Fallback: GET and check magic bytes
        try:
            get_req = _gdrive_req.Request(probe_url, headers=headers)
            with _gdrive_req.urlopen(get_req, timeout=6) as resp:
                ct = resp.headers.get("Content-Type", "").lower()
                chunk = resp.read(16)
            if chunk[:4] == b"%PDF" or chunk[:2] == b"PK":
                return fid, probe_url
        except Exception:
            pass

        return None

    # Parallel probing — use 5 concurrent workers (adjustable)
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(probe_file_id, fid): fid for fid in ids_to_probe}
        for future in as_completed(futures):
            res = future.result()
            if res is not None:
                fid, url = res
                logger.info("Found resume file in folder %s → file ID %s", folder_id, fid)
                # Cancel remaining tasks to save resources
                for f in futures:
                    f.cancel()
                return url

    logger.info("No PDF/DOCX found inside folder %s after probing", folder_id)
    return None


def _fetch_resume_text(url, timeout=12):
    """
    Fetch a Google Drive resume link and extract its text.
    Supports PDF, DOCX, and falls back to HTML scrape.

    If the URL is a Google Drive FOLDER link, crawls the folder to find
    the first PDF or DOCX file inside it, then extracts that.

    Returns extracted text string (empty on failure).
    """
    # ── Folder handling ───────────────────────────────────────────────────────
    if _is_gdrive_folder(url):
        folder_id = _gdrive_folder_id(url)
        if folder_id:
            logger.info("Drive folder detected (%s) — crawling for PDF/DOCX", folder_id)
            file_url = _find_resume_in_gdrive_folder(folder_id, timeout=timeout)
            if file_url:
                # Recurse with the actual file URL (not a folder, safe from infinite loop)
                return _fetch_resume_text(file_url, timeout=timeout)
            else:
                logger.warning("No resume file found in Drive folder %s", folder_id)
                return ""

    # ── Normal file / direct URL handling ────────────────────────────────────
    direct = _gdrive_direct_url(url)
    headers = {"User-Agent": "Mozilla/5.0 (compatible; ResumeGrid/3.0)", "Accept": "*/*"}
    try:
        req = _gdrive_req.Request(direct, headers=headers)
        with _gdrive_req.urlopen(req, timeout=timeout) as resp:
            content_type = resp.headers.get("Content-Type", "").lower()
            raw = resp.read()

        # PDF
        if "pdf" in content_type or raw[:4] == b"%PDF":
            import io
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                parts = [p.extract_text() for p in pdf.pages[:8] if p.extract_text()]
            return "\n".join(parts).strip()

        # DOCX (PK magic bytes = zip = docx)
        if "word" in content_type or "docx" in content_type or raw[:2] == b"PK":
            import io
            return (docx2txt.process(io.BytesIO(raw)) or "").strip()

        # HTML fallback — strip tags
        html = raw.decode("utf-8", errors="replace")
        clean = re.sub(r"<[^>]+>", " ", html)
        clean = re.sub(r"\s+", " ", clean).strip()
        return clean[:6000]

    except Exception as exc:
        logger.warning("Resume fetch failed for %s: %s", url[:80], exc)
        return ""


@app.route("/api/fetch-resume", methods=["POST"])
def api_fetch_resume():
    """
    POST {"url": "https://drive.google.com/..."}
    Returns {"text": "...", "ok": true/false}

    Proxies Google Drive resume downloads to avoid CORS.
    Extracts text from PDF or DOCX automatically.
    """
    body = request.get_json(silent=True) or {}
    url = (body.get("url") or "").strip()

    if not url or not url.startswith("http"):
        return jsonify({"text": "", "ok": False, "error": "invalid url"}), 400

    text = _fetch_resume_text(url)
    logger.info("Resume fetch %s → %d chars", url[:80], len(text))
    return jsonify({"text": text, "ok": bool(text)})


@app.route("/api/rank-candidates", methods=["POST"])
def api_rank_candidates():
    """
    POST {"prompt": "...full ranking prompt with JD + candidate data..."}
    Returns {"scores": [{"id": 0, "score": 72, "reason": "..."}, ...]}

    Uses the same Gemini → Groq → Ollama fallback chain as resume extraction.
    The frontend sends candidates in batches of 20 with their resume text included.
    """
    body = request.get_json(silent=True) or {}
    prompt = (body.get("prompt") or "").strip()

    if not prompt:
        return jsonify({"scores": [], "error": "no prompt"}), 400

    try:
        raw, provider = call_ai(prompt, is_scoring=True)
        logger.info("Rank response from %s: %d chars", provider, len(raw))

        # Strip markdown fences if present
        clean = raw.strip()
        clean = re.sub(r"^```[a-z]*\n?", "", clean)
        clean = re.sub(r"\n?```$", "", clean).strip()

        # Pull out JSON array
        arr_match = re.search(r"\[.*\]", clean, re.DOTALL)
        if arr_match:
            clean = arr_match.group(0)

        scores_raw = json.loads(clean)
        if not isinstance(scores_raw, list):
            raise ValueError("Expected a JSON array")

        scores = []
        for item in scores_raw:
            if isinstance(item, dict) and "id" in item and "score" in item:
                scores.append({
                    "id":     int(item["id"]),
                    "score":  max(0, min(100, int(float(item.get("score", 50))))),
                    "reason": str(item.get("reason", ""))[:300],
                })

        return jsonify({"scores": scores, "provider": provider})

    except json.JSONDecodeError as e:
        logger.error("JSON parse error in rank response: %s | raw: %.300s", e, raw)
        return jsonify({"scores": [], "error": f"JSON parse error: {e}"}), 500
    except Exception as e:
        logger.error("Ranking error: %s", e, exc_info=True)
        return jsonify({"scores": [], "error": str(e)}), 500


@app.route("/api/match-roles", methods=["POST"])
def api_match_roles():
    """
    POST {"jd": "...", "roles": ["Role A", "Role B", ...]}
    Returns {"matched_indices": [0, 2], "reasoning": "..."}

    Uses AI to determine which candidate role buckets are relevant to the given JD.
    Much more accurate than keyword matching — understands semantic role equivalence.
    """
    body = request.get_json(silent=True) or {}
    jd    = (body.get("jd") or "").strip()
    roles = body.get("roles") or []

    if not jd or not roles:
        return jsonify({"matched_indices": [], "reasoning": "missing jd or roles"}), 400

    roles_list = "\n".join(f'{i}: "{r}"' for i, r in enumerate(roles))

    prompt = f"""You are a recruiting assistant. Given a job description and a list of candidate-applied roles, identify which roles from the list are a good match for the JD.

AVAILABLE ROLES (candidates selected one or more of these when applying):
{roles_list}

JOB DESCRIPTION:
{jd[:3000]}

Return ONLY a JSON object:
{{
  "matched_indices": [<list of integer indices>],
  "reasoning": "<one sentence explanation>"
}}

Rules:
- Include a role index if candidates in that role would plausibly be suitable for this JD
- Include multiple indices if the JD genuinely spans multiple domains
- If the JD is for an EIR / GTM / Sales role, do NOT include engineering/developer roles
- If no role matches, return matched_indices as an empty list
- Return ONLY valid JSON, no markdown, no explanation outside the object"""

    try:
        raw, provider = call_ai(prompt, is_scoring=True)
        logger.info("match-roles response from %s", provider)

        clean = raw.strip()
        clean = re.sub(r"^```[a-z]*\n?", "", clean)
        clean = re.sub(r"\n?```$", "", clean).strip()
        obj_match = re.search(r'\{.*\}', clean, re.DOTALL)
        if not obj_match:
            raise ValueError("No JSON object in response")

        parsed = json.loads(obj_match.group())
        matched = [int(i) for i in (parsed.get("matched_indices") or []) if 0 <= int(i) < len(roles)]
        return jsonify({
            "matched_indices": matched,
            "reasoning": str(parsed.get("reasoning", ""))[:200],
            "provider": provider
        })

    except Exception as e:
        logger.error("match-roles error: %s", e, exc_info=True)
        return jsonify({"matched_indices": [], "reasoning": f"AI error: {e}"}), 500

@app.route("/api/download-ranked-excel", methods=["POST"])
def api_download_ranked_excel():
    """
    POST {"candidates": [...], "jd_title": "optional title"}
    Accepts a JSON array of ranked candidate objects (with score + score_reason)
    and returns a formatted .xlsx file as a download.

    The candidates array should be the full candidate objects as stored in the
    frontend (allCandidates after ranking), already sorted by score descending.
    Only candidates that have been scored (score != null) are included.
    """
    body = request.get_json(silent=True) or {}
    candidates = body.get("candidates") or []
    jd_title   = str(body.get("jd_title") or "Ranked").strip()[:60]

    if not candidates:
        return jsonify({"error": "No candidates provided"}), 400

    # Filter to only scored candidates and sort by score desc
    scored = [c for c in candidates if c.get("score") is not None]
    scored.sort(key=lambda c: int(c.get("score") or 0), reverse=True)

    if not scored:
        return jsonify({"error": "No ranked candidates found"}), 400

    logger.info("Generating ranked Excel: %d candidates, JD: %s", len(scored), jd_title)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", start_color="1A3A5C")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font  = Font(name="Arial", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    alt_fill   = PatternFill("solid", start_color="EBF2FA")
    thin       = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
    )

    # ── Column definitions ────────────────────────────────────────────────────
    COLS = [
        ("rank",          "Rank",               8),
        ("score",         "Score (/100)",        10),
        ("score_reason",  "Score Reason",        45),
        ("skill_depth",   "Skill Depth (per JD)", 38),
        ("full_name",     "Full Name",           22),
        ("email",         "Email",               28),
        ("phone",         "Phone",               16),
        ("roles",         "Role(s) Applied",     28),
        ("college",       "College",             28),
        ("degree",        "Degree",              18),
        ("year",          "Year",                10),
        ("duration",      "Duration Available",  18),
        ("start_date",    "Can Start",           15),
        ("location",      "Location",            18),
        ("hyderabad",     "HYD In-Person",       14),
        ("gender",        "Gender",              12),
        ("linkedin",      "LinkedIn",            35),
        ("resume",        "Resume Link",         35),
        ("achievements",  "Achievements",        40),
        ("why_alfaleus",  "Why Alfaleus",        45),
        ("timestamp",     "Submitted",           18),
    ]

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (_, label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, cand in enumerate(scored, 2):
        rank = row_idx - 1
        score_val = cand.get("score")

        for col_idx, (key, _, _) in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if key == "rank":
                cell.value = rank
                cell.font  = Font(name="Arial", size=9, bold=True)

            elif key == "score":
                try:
                    sv = int(score_val)
                    cell.value = sv
                    if sv >= 75:
                        cell.fill = PatternFill("solid", start_color="C6EFCE")
                        cell.font = Font(name="Arial", size=9, bold=True, color="276221")
                    elif sv >= 50:
                        cell.fill = PatternFill("solid", start_color="FFEB9C")
                        cell.font = Font(name="Arial", size=9, bold=True, color="9C6500")
                    else:
                        cell.fill = PatternFill("solid", start_color="FFC7CE")
                        cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")
                except (ValueError, TypeError):
                    cell.value = score_val
                    cell.font  = cell_font

            elif key == "skill_depth":
                depth = cand.get("skill_depth")
                if isinstance(depth, dict):
                    # Format as clean text: "Expert: Python, Django | Intermediate: Docker | Beginner: Redis"
                    order = ["expert", "intermediate", "beginner", "not found"]
                    parts = []
                    for level in order:
                        skills = [k for k, v in depth.items() if str(v).lower() == level]
                        if skills:
                            parts.append(f"{level.title()}: {', '.join(skills)}")
                    cell.value = " | ".join(parts) if parts else ""
                elif isinstance(depth, str):
                    # Already a formatted string — strip star emojis for cleaner Excel
                    import re as _re
                    clean = _re.sub(r'[★☆✗]+ ?', '', depth).strip()
                    cell.value = clean
                else:
                    cell.value = ""
                cell.font = cell_font

            elif key == "roles":
                roles = cand.get("roles") or []
                cell.value = ", ".join(roles) if isinstance(roles, list) else str(roles or "")
                cell.font  = cell_font

            elif key == "hyderabad":
                hyd = cand.get("hyderabad_bool") or cand.get("hyderabad") or ""
                if hyd is True or str(hyd).lower() in ("yes", "true", "1"):
                    cell.value = "Yes"
                    cell.fill  = PatternFill("solid", start_color="C6EFCE")
                    cell.font  = Font(name="Arial", size=9, color="276221")
                else:
                    cell.value = "No"
                    cell.font  = cell_font

            elif key in ("linkedin", "resume"):
                url = str(cand.get(key) or "").strip()
                if url.startswith("http"):
                    cell.value     = url
                    cell.hyperlink = url
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                else:
                    cell.value = url
                    cell.font  = cell_font

            else:
                val = cand.get(key) or ""
                cell.value = str(val) if val else ""
                cell.font  = cell_font

            cell.alignment = cell_align
            cell.border    = thin
            if row_idx % 2 == 0 and key not in ("score", "hyderabad", "linkedin", "resume", "rank"):
                try:
                    if cell.fill.fill_type == "none" or cell.fill.start_color.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = alt_fill
                except Exception:
                    pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Save & return ─────────────────────────────────────────────────────────
    safe_title  = re.sub(r'[^\w\s-]', '', jd_title).strip().replace(' ', '_')
    filename    = f"ranked_{safe_title}_{len(scored)}_candidates.xlsx"
    output_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(output_path)
    logger.info("Ranked Excel saved: %s (%d rows)", filename, len(scored))

    return send_file(
        output_path, as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ── WhatsApp proxy routes (lazy Node process management) ──────────────────────
#
#  The Node/Chromium service is NOT started at container boot.
#  It is spawned on-demand when the user calls /wa/connect and is
#  destroyed automatically by server.js after WA_IDLE_TIMEOUT_MS of
#  inactivity (default 10 min).  Flask only manages the Node *process*
#  lifetime here — Chromium lifecycle is still owned by server.js.

import urllib.request as _urllib_req

WA_SERVICE_URL = os.environ.get("WA_SERVICE_URL", "http://localhost:3001")
WA_PORT        = int(os.environ.get("WA_PORT", "3001"))

_wa_proc: subprocess.Popen | None = None   # the running Node process (or None)


def _wa_process_alive() -> bool:
    """Return True if our Node process is still running."""
    return _wa_proc is not None and _wa_proc.poll() is None


def _spawn_wa_service():
    """
    Start the Node WhatsApp service in the background (non-blocking).
    Does nothing if the process is already running.
    Returns (ok: bool, message: str).
    """
    global _wa_proc

    if _wa_process_alive():
        logger.debug("[WA] spawn requested but process already running (PID %d)", _wa_proc.pid)
        return True, "already running"

    node_script = "/app/whatsapp-service/server.js"
    if not os.path.exists(node_script):
        # Local dev fallback
        node_script = os.path.join(os.path.dirname(__file__), "whatsapp-service", "server.js")

    if not os.path.exists(node_script):
        return False, f"server.js not found at {node_script}"

    try:
        _wa_proc = subprocess.Popen(
            [
                "node",
                "--max-old-space-size=192",
                "--expose-gc",
                "--gc-interval=100",
                node_script,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        logger.info("[WA] Node service spawned (PID %d)", _wa_proc.pid)
    except FileNotFoundError:
        return False, "node binary not found — is Node.js installed?"
    except Exception as exc:
        return False, str(exc)

    # Wait up to 10 s for the HTTP server to be reachable
    health_url = f"http://localhost:{WA_PORT}/health"
    for attempt in range(10):
        time.sleep(1)
        if not _wa_process_alive():
            return False, "Node process exited immediately after spawn"
        try:
            with _urllib_req.urlopen(health_url, timeout=2) as r:
                if r.status == 200:
                    logger.info("[WA] Service healthy after %ds", attempt + 1)
                    return True, "started"
        except Exception:
            pass

    # Process is alive but HTTP isn't answering — still return ok so
    # whatsapp-web.js can continue initialising (QR gen takes a moment)
    if _wa_process_alive():
        logger.warning("[WA] HTTP not ready yet but process is alive — continuing")
        return True, "starting"

    return False, "Node service failed to become healthy in time"


def _wa_request(method, path, body=None):
    """Simple wrapper to call the Node WhatsApp service."""
    url = f"{WA_SERVICE_URL}{path}"
    data = json.dumps(body).encode() if body else None
    headers = {"Content-Type": "application/json"}
    req = _urllib_req.Request(url, data=data, headers=headers, method=method)
    try:
        with _urllib_req.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read()), resp.status
    except _urllib_req.HTTPError as e:
        try:
            return json.loads(e.read()), e.code
        except Exception:
            return {"ok": False, "error": str(e)}, e.code
    except Exception as exc:
        return {"ok": False, "error": str(exc)}, 503


@app.route("/wa/status")
def wa_status():
    # If the Node process isn't running, report disconnected without trying to contact it
    if not _wa_process_alive():
        return jsonify({"status": "disconnected", "qr": None, "phone": None}), 200
    data, code = _wa_request("GET", "/status")
    return jsonify(data), code


@app.route("/wa/connect", methods=["POST"])
def wa_connect():
    """
    Lazily spawns the Node service the first time the user wants to connect.
    Subsequent calls while the process is alive are forwarded straight through.
    """
    if not _wa_process_alive():
        ok, msg = _spawn_wa_service()
        if not ok:
            logger.error("[WA] Failed to spawn Node service: %s", msg)
            return jsonify({"ok": False, "error": f"Could not start WhatsApp service: {msg}"}), 500
        logger.info("[WA] Node service ready — forwarding /connect: %s", msg)

    data, code = _wa_request("POST", "/connect")
    return jsonify(data), code


@app.route("/wa/disconnect", methods=["POST"])
def wa_disconnect():
    if not _wa_process_alive():
        return jsonify({"ok": True, "message": "Not running"}), 200
    data, code = _wa_request("POST", "/disconnect")
    return jsonify(data), code


@app.route("/wa/send", methods=["POST"])
def wa_send():
    if not _wa_process_alive():
        return jsonify({"ok": False, "error": "WhatsApp service not running. Connect first."}), 503
    body = request.get_json(silent=True) or {}
    data, code = _wa_request("POST", "/send", body)
    return jsonify(data), code


@app.route("/wa/send-bulk", methods=["POST"])
def wa_send_bulk():
    if not _wa_process_alive():
        return jsonify({"ok": False, "error": "WhatsApp service not running. Connect first."}), 503
    body = request.get_json(silent=True) or {}
    data, code = _wa_request("POST", "/send-bulk", body)
    return jsonify(data), code


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info("Starting Flask dev server on port %d", port)
    app.run(host="0.0.0.0", port=port, debug=False)